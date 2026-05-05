"""
Utilidades para importar y exportar datos en XLSX
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
from models.producto import Producto
from models.venta import Venta
import os


class ExcelUtils:
    """Utilidades para trabajar con archivos Excel"""
    
    @staticmethod
    def exportar_productos(ruta_archivo):
        """Exporta todos los productos a un archivo XLSX"""
        try:
            productos = Producto.obtener_todos()
            
            # Crear un nuevo workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Productos"
            
            # Definir estilos
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Encabezados
            headers = ["ID", "Código", "Nombre", "Descripción", "Precio", "Stock", "Fecha Creación"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            
            # Datos
            for row, producto in enumerate(productos, start=2):
                ws.cell(row=row, column=1).value = producto.id
                ws.cell(row=row, column=2).value = producto.codigo
                ws.cell(row=row, column=3).value = producto.nombre
                ws.cell(row=row, column=4).value = producto.descripcion
                ws.cell(row=row, column=5).value = producto.precio
                ws.cell(row=row, column=6).value = producto.cantidad_stock
                ws.cell(row=row, column=7).value = producto.fecha_creacion
                
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = border
                    if col == 5:  # Precio
                        ws.cell(row=row, column=col).number_format = '$#,##0.00'
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 20
            
            wb.save(ruta_archivo)
            return {"exito": True, "mensaje": f"Productos exportados a {ruta_archivo}"}
        except Exception as e:
            return {"exito": False, "mensaje": f"Error al exportar productos: {str(e)}"}
    
    @staticmethod
    def exportar_ventas(ruta_archivo):
        """Exporta todas las ventas a un archivo XLSX"""
        try:
            ventas = Venta.obtener_todas()
            
            # Crear un nuevo workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Ventas"
            
            # Definir estilos
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Encabezados
            headers = ["ID", "Código Producto", "Nombre Producto", "Cantidad", "Precio Unitario", "Subtotal", "Fecha Venta"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            
            # Datos
            total = 0
            for row, venta in enumerate(ventas, start=2):
                ws.cell(row=row, column=1).value = venta['id']
                ws.cell(row=row, column=2).value = venta['codigo_producto']
                ws.cell(row=row, column=3).value = venta['nombre_producto']
                ws.cell(row=row, column=4).value = venta['cantidad']
                ws.cell(row=row, column=5).value = venta['precio_unitario']
                ws.cell(row=row, column=6).value = venta['subtotal']
                ws.cell(row=row, column=7).value = venta['fecha_venta']
                
                total += venta['subtotal']
                
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = border
                    if col in [5, 6]:  # Precios
                        ws.cell(row=row, column=col).number_format = '$#,##0.00'
            
            # Agregar fila de total
            total_row = len(ventas) + 2
            ws.cell(row=total_row, column=5).value = "TOTAL:"
            ws.cell(row=total_row, column=5).font = Font(bold=True)
            ws.cell(row=total_row, column=5).alignment = Alignment(horizontal="right")
            ws.cell(row=total_row, column=6).value = total
            ws.cell(row=total_row, column=6).font = Font(bold=True)
            ws.cell(row=total_row, column=6).number_format = '$#,##0.00'
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 20
            
            wb.save(ruta_archivo)
            return {"exito": True, "mensaje": f"Ventas exportadas a {ruta_archivo}"}
        except Exception as e:
            return {"exito": False, "mensaje": f"Error al exportar ventas: {str(e)}"}
    
    @staticmethod
    def importar_productos(ruta_archivo):
        """Importa productos desde un archivo XLSX"""
        try:
            wb = openpyxl.load_workbook(ruta_archivo)
            ws = wb.active
            
            productos_importados = 0
            errores = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:  # Saltar filas vacías
                    continue
                
                try:
                    codigo = str(row[1]).strip() if row[1] else ""
                    nombre = str(row[2]).strip() if row[2] else ""
                    descripcion = str(row[3]).strip() if row[3] else ""
                    precio = float(row[4]) if row[4] else 0
                    cantidad_stock = int(row[5]) if row[5] else 0
                    
                    if codigo and nombre and precio > 0:
                        Producto.crear(codigo, nombre, precio, cantidad_stock, descripcion)
                        productos_importados += 1
                    else:
                        errores.append(f"Fila {row}: datos incompletos o inválidos")
                except Exception as e:
                    errores.append(f"Fila {row}: {str(e)}")
            
            mensaje = f"Se importaron {productos_importados} productos"
            if errores:
                mensaje += f" con {len(errores)} errores"
            
            return {
                "exito": True,
                "mensaje": mensaje,
                "productos_importados": productos_importados,
                "errores": errores
            }
        except Exception as e:
            return {"exito": False, "mensaje": f"Error al importar productos: {str(e)}"}
